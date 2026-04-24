import os
from datetime import datetime
from io import BytesIO
from typing import Any

import pandas as pd
import streamlit as st

def _safe_datetime_series(value) -> pd.Series:
    if value is None:
        return pd.Series(dtype="datetime64[ns]")
    converted = pd.to_datetime(value, errors="coerce")
    if isinstance(converted, pd.Series):
        return converted
    if isinstance(converted, pd.Index):
        return pd.Series(converted)
    if pd.isna(converted):
        return pd.Series(dtype="datetime64[ns]")
    return pd.Series([converted])






def file_summary(uploaded_file, df: pd.DataFrame | None, required_columns: list[str]):
    if not uploaded_file:
        st.info("No file uploaded yet.")
        return False

    st.caption(f"File: {uploaded_file.name}")
    if df is None:
        st.warning("Could not read this file.")
        return False

    c1, c2, c3 = st.columns(3)
    c1.metric("Rows", len(df))
    c2.metric("Columns", len(df.columns))
    c3.metric("Required", len(required_columns))

    missing = [col for col in required_columns if col not in df.columns]
    if missing:
        st.error(f"Missing required columns: {', '.join(missing)}")
        return False
    st.success("Required columns check passed.")
    return True






def export_to_excel(
    df: pd.DataFrame, 
    sheet_name: str = "Detailed Data", 
    additional_sheets: dict[str, pd.DataFrame] = None,
    report_title: str = "Business Intelligence Report",
    summary_metrics: dict[str, Any] = None
) -> bytes:
    """High-fidelity Excel export with professional styling, multi-sheet support, and summary insights."""
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    output = BytesIO()
    
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        # 1. Summary Sheet (If metrics provided)
        if summary_metrics:
            summary_df = pd.DataFrame([
                {"Metric": k, "Value": v} for k, v in summary_metrics.items()
            ])
            summary_df.to_excel(writer, index=False, sheet_name="Executive Summary")
        
        # 2. Main Data Sheet
        df.to_excel(writer, index=False, sheet_name=sheet_name)
        
        # 3. Additional Sheets
        if additional_sheets:
            for s_name, s_df in additional_sheets.items():
                s_df.to_excel(writer, index=False, sheet_name=s_name)
        
        wb = writer.book
        
        # Consistent Styling
        header_fill = PatternFill(start_color='1E293B', end_color='1E293B', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF', size=11)
        summary_fill = PatternFill(start_color='F8FAFC', end_color='F8FAFC', fill_type='solid')
        thin_border = Border(
            left=Side(style='thin', color='CBD5E1'), 
            right=Side(style='thin', color='CBD5E1'), 
            top=Side(style='thin', color='CBD5E1'), 
            bottom=Side(style='thin', color='CBD5E1')
        )
        
        for ws in wb.worksheets:
            is_summary = ws.title == "Executive Summary"
            
            # Header Styling
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = thin_border
            
            # Freeze Header (except summary)
            if not is_summary:
                ws.freeze_panes = 'A2'
            
            # Auto-adjust column widths and apply borders
            for col_idx, col in enumerate(ws.columns, 1):
                max_length = 0
                column_letter = get_column_letter(col_idx)
                
                for cell in col:
                    # Styling every cell
                    cell.border = thin_border
                    if is_summary and cell.row > 1:
                        cell.fill = summary_fill
                        
                    try:
                        val_str = str(cell.value)
                        if len(val_str) > max_length:
                            max_length = len(val_str)
                    except:
                        pass
                
                ws.column_dimensions[column_letter].width = max(min(max_length + 3, 60), 12)

    output.seek(0)
    return output.read()

    output.seek(0)
    return output.read()


def show_last_updated(path: str):
    if not os.path.exists(path):
        return
    updated = datetime.fromtimestamp(os.path.getmtime(path)).strftime("%Y-%m-%d %H:%M:%S")
    st.caption(f"Last updated: {updated}")
