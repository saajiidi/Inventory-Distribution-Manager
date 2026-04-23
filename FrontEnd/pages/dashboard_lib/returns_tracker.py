"""Returns Insights - Dashboard Page.

Premium Streamlit UI for tracking returns, partials, exchanges,
and calculating Net Sales intelligence.
"""

from __future__ import annotations
import time
from datetime import datetime, date, timedelta
from threading import Thread

import streamlit as st
import pandas as pd
import plotly.express as px
import plotly.graph_objects as go
from streamlit_autorefresh import st_autorefresh
from streamlit.runtime.scriptrunner import add_script_run_context

from BackEnd.services.returns_tracker import (
    load_returns_data,
    get_current_sync_window,
    calculate_net_sales_metrics,
    get_issue_type_color,
    DEFAULT_SHEET_URL,
    get_order_items_breakdown,
    track_reordering_customers,
)
from FrontEnd.components import ui
from FrontEnd.utils.error_handler import log_error
from BackEnd.core.logging_config import get_logger

logger = get_logger("returns_tracker_page")


def _load_returns_async(sync_window: str, sales_df_full: pd.DataFrame):
    """Background loader for returns data."""
    try:
        st.session_state["returns_load_started"] = time.time()
        
        # Load the data (blocks this thread, not the UI)
        df_returns = load_returns_data(sync_window=sync_window, sales_df=sales_df_full)
        
        # Store in session state
        st.session_state["returns_data"] = df_returns
        st.session_state["last_returns_sync"] = sync_window
        st.session_state["returns_load_complete"] = True
    except Exception as e:
        log_error(e, context="Returns Background Load")
    finally:
        st.session_state["returns_loading"] = False


def _get_date_range_from_window() -> tuple[date, date]:
    """Get start and end dates from global time window."""
    today = date.today()
    window = st.session_state.get("time_window", "Last Month")

    if window == "MTD":
        start_dt = today.replace(day=1)
        end_dt = today
    elif window == "YTD":
        start_dt = today.replace(month=1, day=1)
        end_dt = today
    elif window == "Last Month":
        end_dt = today
        start_dt = today - timedelta(days=30)
    elif window == "Last 7 Days":
        end_dt = today
        start_dt = today - timedelta(days=7)
    elif window == "Custom Date Range":
        start_dt = st.session_state.get("wc_sync_start_date", today - timedelta(days=30))
        end_dt = st.session_state.get("wc_sync_end_date", today)
    else:
        end_dt = today
        start_dt = today - timedelta(days=30)

    return start_dt, end_dt


def _filter_sales_by_date_range(sales_df: pd.DataFrame, start_dt: date, end_dt: date) -> pd.DataFrame:
    """Filter sales data by date range."""
    if sales_df.empty or "order_date" not in sales_df.columns:
        return sales_df

    # Convert order_date to date if needed
    sales_df = sales_df.copy()
    if pd.api.types.is_datetime64_any_dtype(sales_df["order_date"]):
        mask = (sales_df["order_date"].dt.date >= start_dt) & (sales_df["order_date"].dt.date <= end_dt)
    else:
        # Try to parse if string
        sales_df["_date_parsed"] = pd.to_datetime(sales_df["order_date"], errors="coerce")
        mask = (sales_df["_date_parsed"].dt.date >= start_dt) & (sales_df["_date_parsed"].dt.date <= end_dt)
        sales_df = sales_df.drop(columns=["_date_parsed"])

    return sales_df[mask]


def render_returns_tracker_page() -> None:
    """Main entry point for the Returns Insights page with staged loading."""
    
    # ── Initialize State ──
    if "returns_loading" not in st.session_state:
        st.session_state["returns_loading"] = False
    if "returns_load_complete" not in st.session_state:
        st.session_state["returns_load_complete"] = False

    c1, c2 = st.columns([3, 1])
    with c1:
        st.markdown("### 🔄 Returns Insights")
    with c2:
        if st.button("🔄 Force Refresh", use_container_width=True):
            # Clear cache and state to trigger a fresh background sync
            load_returns_data.clear()
            st.session_state.pop("returns_data", None)
            st.session_state.pop("last_returns_sync", None)
            st.session_state["returns_loading"] = False
            st.session_state["returns_load_complete"] = False
            st.rerun()

    # ── Get Global Time Window ──
    start_dt, end_dt = _get_date_range_from_window()

    # ── Auto Data Sync ──
    sales_df_full = _get_gross_sales_context()
    sales_df = _filter_sales_by_date_range(sales_df_full, start_dt, end_dt)
    sync_window = get_current_sync_window()
    
    # Trigger background load if needed
    needs_load = "returns_data" not in st.session_state or st.session_state.get("last_returns_sync") != sync_window
    is_loading = st.session_state.get("returns_loading", False)

    if needs_load and not is_loading:
        # Start background thread
        st.session_state["returns_loading"] = True
        thread = Thread(target=_load_returns_async, args=(sync_window, sales_df_full), daemon=True)
        add_script_run_context(thread)
        thread.start()
        st.rerun()

    # ── Staged Loading Display ──
    if st.session_state.get("returns_loading", False) and "returns_data" not in st.session_state:
        # Phase 1: Show skeletons while data is being prepared
        st.info("📊 Syncing delivery-issue data from Google Sheets in the background...")
        ui.skeleton_row(count=4)
        st.markdown("<br>", unsafe_allow_html=True)
        ui.skeleton_row(count=3)
        
        # Poll for completion
        st_autorefresh(interval=3000, key="returns_sync_refresh")
        return

    # Data is available (either from this sync or previous one)
    if "returns_data" not in st.session_state or st.session_state.returns_data.empty:
        st.info("📊 No Returns Data available. Check the source connection.")
        return

    # If still loading but we have OLD data, show a small indicator
    if st.session_state.get("returns_loading", False):
        st.caption("🔄 Data is refreshing in the background... showing cached snapshot.")
        st_autorefresh(interval=5000, key="returns_background_refresh")

    df = st.session_state.returns_data.copy()

    # ── Date Range Filter ──
    df = _render_date_filter(df)

    # ── Calculate Total Items Sold for % Calculations ──
    total_items_sold = 0
    if not sales_df.empty and "quantity" in sales_df.columns:
        total_items_sold = int(sales_df["quantity"].sum())

    # ── Compute Metrics ──
    metrics = calculate_net_sales_metrics(df, sales_df=sales_df, total_items_sold=total_items_sold)

    # ── TABS ──
    tab_dash, tab_recovery, tab_inventory, tab_ledger = st.tabs([
        "📊 Executive Dashboard", 
        "🛡️ Recovery & Loyalty", 
        "📦 Return Inventory", 
        "📋 Detailed Ledger"
    ])

    with tab_dash:
        # ── KPI Cards ──
        _render_kpi_cards(metrics)
        _render_financial_impact_summary(metrics)

        if df.empty:
            st.info("No returns logged within this specific time frame.")
        else:
            # ── Charts ──
            st.markdown("---")
            _render_charts(df, metrics, sales_df)
            
            # ── Export ──
            st.markdown("---")
            _render_export(df, metrics)

    with tab_recovery:
        _render_customer_recovery(df, sales_df)

    with tab_inventory:
        _render_return_inventory(df, sales_df)

    with tab_ledger:
        _render_details_table(df, sales_df)



# ═══════════════════════════════════════════════════════════════════
# DATE FILTER
# ═══════════════════════════════════════════════════════════════════

def _render_date_filter(df: pd.DataFrame) -> pd.DataFrame:
    """Apply global date range filter mapping from Business Intelligence dashboard window."""
    today = date.today()
    window = st.session_state.get("time_window", "Last Month")

    start_dt = end_dt = today
    if window == "MTD":
        start_dt = today.replace(day=1)
    elif window == "YTD":
        start_dt = today.replace(month=1, day=1)
    elif window == "Custom Date Range":
        start_dt = st.session_state.get("wc_sync_start_date", today - timedelta(days=30))
        end_dt = st.session_state.get("wc_sync_end_date", today)
    else:
        window_map = {
            "Last Day": 1,
            "Last 3 Days": 3,
            "Last 4 Days": 4,
            "Last 7 Days": 7,
            "Last 15 Days": 15,
            "Last Month": 30,
            "Last 3 Months": 90,
            "Last Quarter": 90,
            "Last Half Year": 180,
            "Last 9 Months": 270,
            "Last Year": 365
        }
        days_back = window_map.get(window, 30)
        start_dt = today - timedelta(days=days_back)

    # Issue type filter - only if column exists
    if "issue_type" in df.columns:
        all_types = sorted(df["issue_type"].unique().tolist())
        selected_types = st.multiselect(
            "Issue Types", all_types,
            default=all_types,
            key="returns_type_filter"
        )
    else:
        selected_types = []

    st.caption(f"📅 Using synchronized date range: **{start_dt}** to **{end_dt}** ({window})")

    # Filter dataframe - only if date column exists
    if "date" in df.columns and not df.empty:
        mask = (df["date"].dt.date >= start_dt) & (df["date"].dt.date <= end_dt)
        if selected_types:
            mask &= df["issue_type"].isin(selected_types)
        return df[mask]

    return df


# ═══════════════════════════════════════════════════════════════════
# GROSS SALES CONTEXT
# ═══════════════════════════════════════════════════════════════════

def _get_gross_sales_context():
    """Try to pull sales data for mapping. Prefers full cache for better coverage."""
    from BackEnd.services.hybrid_data_loader import load_cached_woocommerce_history
    
    # 1. Try to load full history from cache first for maximum mapping coverage
    full_cache = load_cached_woocommerce_history()
    if not full_cache.empty:
        return full_cache

    # 2. Fallback to active dashboard data if cache is unavailable
    if "dashboard_data" in st.session_state:
        data = st.session_state.dashboard_data
        return data.get("sales_active", pd.DataFrame())

    return pd.DataFrame()


# ═══════════════════════════════════════════════════════════════════
# KPI CARDS
# ═══════════════════════════════════════════════════════════════════

def _render_kpi_cards(metrics: dict) -> None:
    """Render the executive KPI cards using premium components."""
    st.markdown("#### 📦 Operational Intelligence")
    
    t_ord = metrics.get('total_orders', 0)
    
    def format_pct(val):
        if t_ord > 0:
            return f"{val:,} ({(val / t_ord * 100):.1f}%)"
        return f"{val:,}"

    cols = st.columns(4)
    
    with cols[0]:
        ui.metric_highlight(
            label="Total Issues",
            value=format_pct(metrics.get('total_issues', 0)),
            help_text=f"Out of {t_ord:,} total orders",
            icon="📦"
        )

    with cols[1]:
        ui.metric_highlight(
            label="Returns",
            value=format_pct(metrics.get('return_count', 0)),
            help_text=f"Paid: {metrics.get('paid_return_count', 0)} | Non-Paid: {metrics.get('non_paid_return_count', 0)}",
            icon="🔴"
        )

    with cols[2]:
        ui.metric_highlight(
            label="Partials",
            value=format_pct(metrics.get('partial_count', 0)),
            help_text=f"৳{metrics.get('partial_amounts', 0):,.0f} impact",
            icon="🟡"
        )

    with cols[3]:
        ui.metric_highlight(
            label="Exchanges",
            value=format_pct(metrics.get('exchange_count', 0)),
            help_text="Product/Size swaps",
            icon="🟣"
        )


def _render_financial_impact_summary(metrics: dict) -> None:
    """Render decision-ready financial impact cards using premium components."""
    st.markdown("#### 💰 Financial Integrity & Yield")

    gross = metrics.get('gross_sales', 0)
    net_sales = metrics.get('net_sales', 0)
    net_yield_pct = metrics.get('net_yield_pct', (net_sales / gross * 100) if gross > 0 else 0.0)

    total_ret_qty = metrics.get('total_return_qty_all', 0)
    total_items_sold = metrics.get('total_items_sold', 0)
    total_returned_items_pct = metrics.get('total_returned_items_pct', 0.0)
    returned_orders_pct = metrics.get('returned_orders_pct', 0.0)
    partial_loss = metrics.get('partial_loss', metrics.get('partial_amounts', 0))

    # Primary Row: High-level financial outcome
    c1, c2, c3 = st.columns(3)
    with c1:
        ui.metric_highlight(
            label="Net Settled Sales",
            value=f"৳{net_sales:,.0f}",
            help_text=f"After {metrics.get('return_count', 0)} returns & {metrics.get('partial_count', 0)} partials",
            icon="💰"
        )
    with c2:
        ui.metric_highlight(
            label="Net Revenue Yield",
            value=f"{net_yield_pct:.1f}%",
            help_text=f"Efficiency: ৳{net_sales:,.0f} / ৳{gross:,.0f}",
            icon="📊"
        )
    with c3:
        ui.metric_highlight(
            label="Total Loss Attribution",
            value=f"৳{(metrics.get('return_value_extracted', 0) + partial_loss):,.0f}",
            help_text="Revenue lost to returns and partials",
            icon="📉"
        )

    # Secondary Row: Operational impact
    c4, c5, c6 = st.columns(3)
    with c4:
        items_pct_text = f"{total_returned_items_pct:.1f}% of {total_items_sold:,} units" if total_items_sold > 0 else "0% items returned"
        ui.metric_highlight(
            label="Returned Item Volume",
            value=f"{total_ret_qty} Units",
            help_text=items_pct_text,
            icon="📦"
        )
    with c5:
        ui.metric_highlight(
            label="Returned Order Share",
            value=f"{returned_orders_pct:.1f}%",
            help_text=f"1 in every {int(100/returned_orders_pct) if returned_orders_pct > 0 else 'N/A'} orders",
            icon="📈"
        )
    with c6:
        ui.metric_highlight(
            label="Exchanged Items",
            value=f"{metrics.get('total_exchanged_items', 0)} Units",
            help_text="Product swaps (No revenue loss)",
            icon="🔄"
        )

    # Financial Integrity Chart
    returns_ready = not metrics.get("daily_financials", pd.DataFrame()).empty
    if returns_ready:
        import plotly.graph_objects as go
        fin_plot = metrics.get("daily_financials", pd.DataFrame()).copy()
        fin_plot["date"] = pd.to_datetime(fin_plot["date"], errors="coerce")
        fin_plot = fin_plot.dropna(subset=["date"]).sort_values("date")

        if not fin_plot.empty:
            st.markdown("---")
            fig_gap = go.Figure()

            # Layer 1: Deep Red flame base
            fig_gap.add_trace(go.Scatter(
                x=fin_plot['date'], y=fin_plot['net_sales'] * 0.3,
                fill='tozeroy', mode='lines',
                line=dict(color='rgba(220, 20, 60, 0.3)', width=0),
                fillcolor='rgba(220, 20, 60, 0.4)',
                name='Net Settled', stackgroup='one', showlegend=False
            ))

            # Layer 2: Orange-red
            fig_gap.add_trace(go.Scatter(
                x=fin_plot['date'], y=fin_plot['net_sales'] * 0.6,
                fill='tonexty', mode='lines',
                line=dict(color='rgba(255, 69, 0, 0.4)', width=0),
                fillcolor='rgba(255, 69, 0, 0.5)',
                name='Net Settled', stackgroup='one', showlegend=False
            ))

            # Layer 3: Flame Orange
            fig_gap.add_trace(go.Scatter(
                x=fin_plot['date'], y=fin_plot['net_sales'],
                fill='tonexty', mode='lines',
                line=dict(color='rgba(255, 140, 0, 1.0)', width=3),
                fillcolor='rgba(255, 140, 0, 0.6)',
                name='Net Settled', stackgroup='one'
            ))

            # Layer 4: Golden Yellow
            fig_gap.add_trace(go.Scatter(
                x=fin_plot['date'], y=fin_plot['gross_sales'],
                fill='tonexty', mode='lines',
                line=dict(color='rgba(255, 215, 0, 1.0)', width=2),
                fillcolor='rgba(255, 215, 0, 0.4)',
                name='Gross Verified', stackgroup='one'
            ))

            fig_gap.update_layout(
                height=280,
                title="Sales Integrity Gap (Gross vs. Net Settled Sales)",
                paper_bgcolor='rgba(0,0,0,0)',
                plot_bgcolor='rgba(0,0,0,0)',
                margin=dict(l=0, r=0, t=40, b=0),
                hovermode="x unified",
                legend=dict(orientation="h", y=1.1, x=0.5, xanchor="center")
            )
            st.plotly_chart(fig_gap, width="stretch")

    confidence = metrics.get("attribution_confidence_pct", 0.0)
    matched_items = metrics.get("matched_returned_items", 0)
    estimated_items = metrics.get("estimated_returned_items", 0)
    st.caption(
        f"Financial attribution confidence: {confidence:.1f}% | "
        f"Matched returned items: {matched_items:,} | Unmatched/estimated items: {estimated_items:,}"
    )




# ═══════════════════════════════════════════════════════════════════
# CHARTS
# ═══════════════════════════════════════════════════════════════════

def _render_charts(df: pd.DataFrame, metrics: dict, sales_df: pd.DataFrame) -> None:
    """Render the analytics charts."""

    if df.empty or "date" not in df.columns:
        st.info("📊 Charts will appear once return data is loaded.")
        return

    tab1, tab2, tab3, tab4, tab5, tab6 = st.tabs([
        "📈 Monthly Trends",
        "🥧 Return Reasons",
        "📦 Product Heatmap",
        "🛡️ Customer Recovery",
        "📋 Return Inventory",
        "📦 Returned Items List",
    ])

    with tab1:
        _render_monthly_trend(df)

    with tab2:
        _render_reason_charts(metrics)

    with tab3:
        _render_product_heatmap(df)

    with tab4:
        _render_customer_recovery(df, sales_df)

    with tab5:
        _render_return_inventory(df, sales_df)

    with tab6:
        _render_returned_items_list(df)


def _render_monthly_trend(df: pd.DataFrame) -> None:
    """Monthly issue count trend with type breakdown."""
    if "date" not in df.columns or df.empty:
        st.info("Not enough data for monthly trends.")
        return

    monthly = (
        df.groupby([pd.Grouper(key="date", freq="ME"), "issue_type"])
        .agg(count=("order_id", "nunique"))
        .reset_index()
    )

    if monthly.empty:
        st.info("Not enough data for monthly trends.")
        return

    monthly["month_label"] = monthly["date"].dt.strftime("%b %Y")

    color_map = {t: get_issue_type_color(t) for t in monthly["issue_type"].unique()}

    fig = px.bar(
        monthly,
        x="month_label", y="count",
        color="issue_type",
        color_discrete_map=color_map,
        title="Monthly Issue Breakdown",
        labels={"count": "Orders", "month_label": "Month", "issue_type": "Type"},
        barmode="stack",
    )
    fig.update_layout(
        plot_bgcolor="rgba(0,0,0,0)",
        paper_bgcolor="rgba(0,0,0,0)",
        font=dict(family="Inter, sans-serif"),
        legend=dict(orientation="h", y=-0.2),
        margin=dict(l=20, r=20, t=50, b=20),
    )
    st.plotly_chart(fig, width="stretch")

    # Total trend line
    total_monthly = (
        df.drop_duplicates(subset=["order_id"])
        .groupby(pd.Grouper(key="date", freq="ME"))
        .size()
        .reset_index(name="total_issues")
    )
    total_monthly["month_label"] = total_monthly["date"].dt.strftime("%b %Y")

    fig2 = go.Figure()
    fig2.add_trace(go.Scatter(
        x=total_monthly["month_label"],
        y=total_monthly["total_issues"],
        mode="lines+markers",
        name="Total Issues",
        line=dict(color="#3b82f6", width=3),
        marker=dict(size=8),
    ))
    
    # Also add Net Returns value if available (just count of actual returns)
    return_mask = df["issue_type"].isin(["Paid Return", "Non Paid Return"])
    return_monthly = (
        df[return_mask].drop_duplicates(subset=["order_id"])
        .groupby(pd.Grouper(key="date", freq="ME"))
        .size()
        .reset_index(name="return_issues")
    )
    if not return_monthly.empty:
        return_monthly["month_label"] = return_monthly["date"].dt.strftime("%b %Y")
        fig2.add_trace(go.Scatter(
            x=return_monthly["month_label"],
            y=return_monthly["return_issues"],
            mode="lines+markers",
            name="True Returns",
            line=dict(color="#ef4444", width=3, dash="dot"),
            marker=dict(size=8),
        ))

    fig2.update_layout(
        title="Issue Velocity Over Time",
        plot_bgcolor="rgba(0,0,0,0)",
        paper_bgcolor="rgba(0,0,0,0)",
        font=dict(family="Inter, sans-serif"),
        margin=dict(l=20, r=20, t=50, b=20),
        yaxis_title="Count",
    )
    st.plotly_chart(fig2, width="stretch")


def _render_reason_charts(metrics: dict) -> None:
    """Return reason distribution charts."""
    reasons = metrics.get("reason_counts", {})

    if not reasons:
        st.info("No return reason data available.")
        return

    c1, c2 = st.columns(2)

    with c1:
        reason_df = pd.DataFrame([
            {"Reason": k, "Count": v} for k, v in reasons.items()
        ]).sort_values("Count", ascending=False)

        fig = px.pie(
            reason_df, values="Count", names="Reason",
            title="Return Reasons Distribution",
            hole=0.45,
            color_discrete_sequence=px.colors.qualitative.Set2,
        )
        fig.update_layout(
            plot_bgcolor="rgba(0,0,0,0)",
            paper_bgcolor="rgba(0,0,0,0)",
            font=dict(family="Inter, sans-serif"),
            margin=dict(l=20, r=20, t=50, b=20),
        )
        st.plotly_chart(fig, width="stretch")

    with c2:
        fig2 = px.bar(
            reason_df,
            x="Count", y="Reason",
            orientation="h",
            title="Return Reasons (Count)",
            color="Count",
            color_continuous_scale="Reds",
        )
        fig2.update_layout(
            plot_bgcolor="rgba(0,0,0,0)",
            paper_bgcolor="rgba(0,0,0,0)",
            font=dict(family="Inter, sans-serif"),
            margin=dict(l=20, r=20, t=50, b=20),
            showlegend=False,
            yaxis=dict(autorange="reversed"),
        )
        st.plotly_chart(fig2, width="stretch")


def _render_product_heatmap(df: pd.DataFrame) -> None:
    """Product-level return frequency heatmap."""
    # Extract product category from product_details
    issue_df = df[df["issue_type"].isin([
        "Paid Return", "Non Paid Return", "Partial", "Exchange"
    ])].copy()

    if issue_df.empty:
        st.info("No product-level data available.")
        return

    # Extract rough category from product_details
    issue_df["product_category"] = issue_df["product_details"].apply(_extract_product_category)
    issue_df = issue_df[issue_df["product_category"] != "Unknown"]

    if issue_df.empty:
        st.info("Could not extract product categories from details.")
        return

    heatmap_data = (
        issue_df
        .groupby(["product_category", "issue_type"])
        .agg(count=("order_id", "nunique"))
        .reset_index()
    )

    pivot = heatmap_data.pivot_table(
        index="product_category", columns="issue_type",
        values="count", fill_value=0, aggfunc="sum"
    )

    # Sort by total descending
    pivot["_total"] = pivot.sum(axis=1)
    pivot = pivot.sort_values("_total", ascending=True).drop(columns=["_total"])

    fig = px.imshow(
        pivot.values,
        x=pivot.columns.tolist(),
        y=pivot.index.tolist(),
        title="Return Frequency by Product Category",
        color_continuous_scale="YlOrRd",
        aspect="auto",
        labels=dict(color="Count"),
    )
    fig.update_layout(
        plot_bgcolor="rgba(0,0,0,0)",
        paper_bgcolor="rgba(0,0,0,0)",
        font=dict(family="Inter, sans-serif"),
        margin=dict(l=20, r=20, t=50, b=20),
        height=max(400, len(pivot) * 30 + 100),
    )
    st.plotly_chart(fig, width="stretch")


def _extract_product_category(details: str) -> str:
    """Extract a rough product category from product details text."""
    if not details or details.strip() == "":
        return "Unknown"

    det = details.lower()

    categories = [
        ("Jeans", ["jeans", "denim pant"]),
        ("Flannel Shirt", ["flannel"]),
        ("Kaftan Shirt", ["kaftan"]),
        ("Contrast Shirt", ["contrast stitch"]),
        ("HS Shirt", ["half shirt", "hs-shirt", "hs shirt", "cotton shirt"]),
        ("FS Shirt", ["full sleeve shirt", "fs shirt", "polka dot", "denim shirt"]),
        ("Polo Shirt", ["polo"]),
        ("T-Shirt", ["tshirt", "t-shirt", "active wear"]),
        ("FS T-Shirt", ["fs-tshirt", "fs tshirt", "full sleeve.*t-shirt",
                         "full sleeve.*tshirt", "fs-t-shirt"]),
        ("Sweatshirt", ["sweatshirt"]),
        ("Turtleneck", ["turtleneck"]),
        ("Panjabi", ["panjabi"]),
        ("Trouser", ["trousers", "trouser"]),
        ("Twill/Chino", ["twill", "chino", "jogger"]),
        ("Wallet", ["wallet"]),
        ("Belt", ["belt"]),
        ("Bag", ["bagpack", "bag"]),
        ("Boxer", ["boxer"]),
        ("Bundle", ["pack", "combo", "deal", "dbb"]),
    ]

    for cat_name, keywords in categories:
        for kw in keywords:
            if kw in det:
                return cat_name

    return "Other"


def _render_customer_recovery(df: pd.DataFrame, sales_df: pd.DataFrame) -> None:
    """Analyze and display customers who reordered after issues."""
    st.markdown("#### 🛡️ Customer Loyalty & Recovery Analysis")
    st.caption("Tracking customers who returned/exchanged products but stayed loyal and ordered again.")
    
    reorders = track_reordering_customers(df, sales_df)
    
    if reorders.empty:
        st.info("No reordering events detected for the selected group of customers yet.")
        return
        
    # Stats
    total_reordered = len(reorders)
    avg_days = reorders["Days to Reorder"].mean()
    
    c1, c2 = st.columns(2)
    with c1:
        ui_metric_small("Successfully Recovered", f"{total_reordered}", "🛡️")
    with c2:
        ui_metric_small("Avg. Recovery Days", f"{avg_days:.1f} days", "⏳")
        
    st.markdown("##### 📋 Recovery Ledger")
    st.dataframe(
        reorders,
        width="stretch",
        hide_index=True,
        column_config={
            "Days to Reorder": st.column_config.NumberColumn(format="%d days"),
        }
    )


def _render_return_inventory(df: pd.DataFrame, sales_df: pd.DataFrame) -> None:
    """Explode order-level returns into an item-centric inventory view."""
    st.markdown("#### 📋 Return Item Inventory")
    st.caption("A granular list of every individual product returned, including size and category.")

    # Guard: Check required columns exist
    if "date" not in df.columns:
        st.info("📊 Return data is loading... Date information not yet available.")
        return

    # 1. Prepare exploded item list
    item_rows = []

    # Filter for returns only
    return_mask = df["issue_type"].isin(["Paid Return", "Non Paid Return", "Partial", "Exchange"])
    return_df = df[return_mask].copy()

    for _, row in return_df.iterrows():
        items = row.get("returned_items", [])
        if not isinstance(items, list): continue
        
        for item in items:
            if not isinstance(item, dict): continue
            
            # Map SKU if missing
            sku = item.get("sku", "N/A")
            if sku == "N/A":
                # Try to resolve SKU now
                name = item.get("name", "")
                order_sales = sales_df[sales_df["order_id"].astype(str) == str(row["order_id"])]
                match = order_sales[order_sales["item_name"].str.contains(name, case=False, na=False, regex=False)]
                if not match.empty:
                    sku = match.iloc[0].get("sku", "N/A")

            item_rows.append({
                "Date": row["date"].strftime("%Y-%m-%d") if pd.notnull(row["date"]) else "N/A",
                "Order ID": row["order_id_raw"],
                "Type": row["issue_type"],
                "Product": item.get("name", "Unknown"),
                "SKU": sku,
                "Size": item.get("size", "N/A"),
                "Qty": item.get("qty", 1),
                "Category": item.get("category", "General"),
                "Reason": row.get("return_reason", "N/A")
            })

    if not item_rows:
        st.info("No individual returned items found in the current selection.")
        return

    item_df = pd.DataFrame(item_rows)

    # 2. Filters for the Inventory
    from BackEnd.core.categories import get_master_category_list, format_category_label
    
    c1, c2, c3 = st.columns(3)
    with c1:
        # Use master category list for consistent hierarchy display
        master_cats = get_master_category_list()
        # Filter to only categories present in data but preserve master order
        available_cats = [c for c in master_cats if c in item_df["Category"].values]
        cat_filter = st.multiselect("Filter Category", options=available_cats, format_func=format_category_label)
    with c2:
        type_filter = st.multiselect("Filter Issue Type", options=sorted(item_df["Type"].unique()))
    with c3:
        search_query = st.text_input("🔍 Search Product/SKU", placeholder="Enter name or SKU...")

    # Apply filters
    if cat_filter:
        item_df = item_df[item_df["Category"].isin(cat_filter)]
    if type_filter:
        item_df = item_df[item_df["Type"].isin(type_filter)]
    if search_query:
        item_df = item_df[
            item_df["Product"].str.contains(search_query, case=False, na=False) |
            item_df["SKU"].str.contains(search_query, case=False, na=False)
        ]

    # 3. Render
    st.dataframe(
        item_df,
        width="stretch",
        hide_index=True,
        column_config={
            "Order ID": st.column_config.TextColumn("Order ID"),
            "Date": st.column_config.DateColumn("Date"),
            "Qty": st.column_config.NumberColumn("Qty", format="%d"),
        }
    )

    # 4. Export Small
    csv = item_df.to_csv(index=False).encode('utf-8')
    st.download_button(
        label="📥 Download Return Inventory CSV",
        data=csv,
        file_name=f"return_inventory_{datetime.now().strftime('%Y%m%d')}.csv",
        mime='text/csv',
    )


def _render_returned_items_list(df: pd.DataFrame) -> None:
    """Display a detailed list of returned items from all Returns.

    This view shows individual items returned from the 'Issue Or Product Details' column,
    including Paid Return, Non Paid Return, and Partial.
    """
    st.markdown("#### 📦 Returned Items Detail List")
    st.caption("Individual items from ALL RETURNS (Paid, Non Paid, Partial). Data extracted from 'Issue Or Product Details' column.")

    # Guard: Check required columns exist
    if "date" not in df.columns:
        st.info("📊 Return data is loading... Date information not yet available.")
        return

    # Debug info
    with st.expander("🔍 Debug Data Info", expanded=False):
        st.write(f"Total rows in dataset: {len(df)}")
        st.write(f"Issue types: {df['issue_type'].value_counts().to_dict() if not df.empty else 'N/A'}")
        if not df.empty and 'returned_items' in df.columns:
            sample = df[df['returned_items'].apply(lambda x: len(x) > 0 if isinstance(x, list) else False)].head(3)
            st.write("Sample rows with items:")
            st.dataframe(sample[['order_id_raw', 'issue_type', 'product_details', 'returned_items']])

    # Filter for Returns (Paid, Non Paid, and Partial)
    return_mask = df["issue_type"].isin(["Paid Return", "Non Paid Return", "Partial"])
    returns_df = df[return_mask].copy()

    if returns_df.empty:
        st.warning("⚠️ No Return items found in the current selection.")
        st.info("Check the Debug Data Info above to see what issue types are available.")
        return

    # Explode items from returned_items column
    item_rows = []
    for _, row in returns_df.iterrows():
        items = row.get("returned_items", [])
        if not isinstance(items, list):
            continue

        for item in items:
            if not isinstance(item, dict):
                continue

            item_rows.append({
                "Date": row["date"].strftime("%Y-%m-%d") if pd.notnull(row["date"]) else "N/A",
                "Order ID": row["order_id_raw"],
                "Order Number": row.get("order_id", "N/A"),
                "Issue Type": row.get("issue_type", "N/A"),
                "SKU": item.get("sku", "N/A"),
                "WC Matched": "✅" if item.get("matched_from_wc") else "❌",
                "Product Name": item.get("name", "Unknown"),
                "Size": item.get("size", "N/A"),
                "Qty": item.get("qty", 1),
                "Price": item.get("price", 0),
                "Revenue Impact": item.get("revenue_impact", 0),
                "Category": item.get("category", "General"),
                "Return Reason": row.get("return_reason", "N/A"),
                "Product Details": row.get("product_details", "")[:100] + "..." if len(row.get("product_details", "")) > 100 else row.get("product_details", ""),
            })

    if not item_rows:
        st.info("No individual items found in Returns.")
        return

    items_df = pd.DataFrame(item_rows)

    # Summary metrics
    total_items = len(items_df)
    total_qty = items_df["Qty"].sum()
    unique_orders = items_df["Order ID"].nunique()

    # Summary row
    m1, m2, m3, m4 = st.columns(4)
    with m1:
        ui_metric_small("Total Items", f"{total_items}", "📦")
    with m2:
        ui_metric_small("Total Qty", f"{total_qty}", "🔢")
    with m3:
        ui_metric_small("Unique Orders", f"{unique_orders}", "📋")
    with m4:
        unique_skus = items_df[items_df["SKU"] != "N/A"]["SKU"].nunique()
        ui_metric_small("Unique SKUs", f"{unique_skus}", "🏷️")

    st.markdown("---")

    # Filters
    c1, c2, c3, c4 = st.columns(4)
    with c1:
        from BackEnd.core.categories import get_master_category_list, format_category_label
        master_cats = get_master_category_list()
        available_cats = [c for c in master_cats if c in items_df["Category"].values]
        cat_filter = st.multiselect("Filter Category", options=available_cats, format_func=format_category_label, key="returned_items_cat")
    with c2:
        search_product = st.text_input("🔍 Search Product", placeholder="Enter product name...", key="returned_items_product")
    with c3:
        search_sku = st.text_input("🔍 Search SKU", placeholder="Enter SKU...", key="returned_items_sku")
    with c4:
        search_order = st.text_input("🔍 Search Order ID", placeholder="Enter order ID...", key="returned_items_order")

    # Apply filters
    filtered_df = items_df.copy()
    if cat_filter:
        filtered_df = filtered_df[filtered_df["Category"].isin(cat_filter)]
    if search_product:
        filtered_df = filtered_df[filtered_df["Product Name"].str.contains(search_product, case=False, na=False)]
    if search_sku:
        filtered_df = filtered_df[filtered_df["SKU"].astype(str).str.contains(search_sku, case=False, na=False)]
    if search_order:
        filtered_df = filtered_df[filtered_df["Order ID"].astype(str).str.contains(search_order, case=False, na=False)]

    # Display data table
    st.dataframe(
        filtered_df,
        width="stretch",
        hide_index=True,
        column_config={
            "Date": st.column_config.DateColumn("Date"),
            "Order ID": st.column_config.TextColumn("Order ID"),
            "Order Number": st.column_config.TextColumn("Order Number"),
            "Issue Type": st.column_config.TextColumn("Issue Type"),
            "SKU": st.column_config.TextColumn("SKU"),
            "WC Matched": st.column_config.TextColumn("WC Matched"),
            "Product Name": st.column_config.TextColumn("Product Name"),
            "Size": st.column_config.TextColumn("Size"),
            "Qty": st.column_config.NumberColumn("Qty", format="%d"),
            "Price": st.column_config.NumberColumn("Price", format="৳%.0f"),
            "Revenue Impact": st.column_config.NumberColumn("Revenue Impact", format="৳%.0f"),
            "Category": st.column_config.TextColumn("Category"),
            "Return Reason": st.column_config.TextColumn("Return Reason"),
            "Product Details": st.column_config.TextColumn("Raw Details", width="large"),
        }
    )

    st.caption(f"Showing {len(filtered_df)} returned items from {unique_orders} Return orders (Paid, Non Paid, Partial)")

    # Export option
    csv = filtered_df.to_csv(index=False).encode('utf-8')
    st.download_button(
        label="📥 Download Returned Items CSV",
        data=csv,
        file_name=f"returned_items_{datetime.now().strftime('%Y%m%d')}.csv",
        mime='text/csv',
        key="returned_items_csv"
    )

    # ── RETURN REASON ANALYSIS ──
    st.markdown("---")
    st.markdown("#### 🔍 Why Are Items Being Returned?")
    st.caption("Analysis of return reasons to predict patterns")

    if not filtered_df.empty and 'Return Reason' in filtered_df.columns:
        reason_counts = filtered_df['Return Reason'].value_counts()
        if not reason_counts.empty:
            c1, c2 = st.columns(2)
            with c1:
                st.markdown("**📊 Return Reason Breakdown:**")
                reason_df = reason_counts.reset_index()
                reason_df.columns = ['Reason', 'Count']
                reason_df['% of Returns'] = (reason_df['Count'] / reason_df['Count'].sum() * 100).round(1)
                st.dataframe(reason_df, width="stretch", hide_index=True)

            with c2:
                st.markdown("**💡 Predictions & Insights:**")
                top_reason = reason_counts.index[0]
                top_pct = reason_counts.iloc[0] / reason_counts.sum() * 100

                insights = []
                if top_reason == "Size Issue":
                    insights.append(f"🔴 **{top_pct:.1f}% Size Issues** - Consider adding size charts or fit guides")
                elif top_reason == "Quality Issue":
                    insights.append(f"🔴 **{top_pct:.1f}% Quality Issues** - Review supplier quality control")
                elif top_reason == "Color Issue":
                    insights.append(f"🔴 **{top_pct:.1f}% Color Issues** - Improve product photography accuracy")
                elif top_reason == "CNR":
                    insights.append(f"🔴 **{top_pct:.1f}% CNR** - Enhance delivery communication/scheduling")
                elif top_reason == "Changed Mind":
                    insights.append(f"🟡 **{top_pct:.1f}% Changed Mind** - Consider better product descriptions")
                else:
                    insights.append(f"🟡 **{top_pct:.1f}% {top_reason}** - Review this category")

                # Size-related insights
                if 'Size' in filtered_df.columns:
                    size_issues = filtered_df[filtered_df['Return Reason'] == 'Size Issue']
                    if not size_issues.empty:
                        common_sizes = size_issues['Size'].value_counts().head(3)
                        if not common_sizes.empty:
                            insights.append(f"📏 Most returned sizes: {', '.join(common_sizes.index)}")

                # Category insights
                cat_issues = filtered_df['Category'].value_counts().head(3)
                if not cat_issues.empty:
                    insights.append(f"👕 Top returned categories: {', '.join(cat_issues.index)}")

                for insight in insights:
                    st.markdown(f"- {insight}")
        else:
            st.info("No return reason data available for analysis.")
    else:
        st.info("Insufficient data for return reason prediction.")


def ui_metric_small(label: str, value: str, icon: str):
    """Small metric helper with premium glassmorphism styling."""
    st.markdown(f"""
        <div style="
            background: var(--surface);
            border: 1px solid var(--outline);
            border-radius: 12px;
            padding: 12px;
            text-align: center;
            margin-bottom: 10px;
            box-shadow: 0 4px 12px rgba(0,0,0,0.05);
            backdrop-filter: blur(10px);
        ">
            <div style="font-size: 1.2rem; margin-bottom: 4px;">{icon}</div>
            <div style="font-size: 0.7rem; color: #6b7280; text-transform: uppercase; font-weight: 600; letter-spacing: 0.5px;">{label}</div>
            <div style="font-size: 1.3rem; font-weight: 700; color: var(--text-color);">{value}</div>
        </div>
    """, unsafe_allow_html=True)


# ═══════════════════════════════════════════════════════════════════
# DETAILS TABLE
# ═══════════════════════════════════════════════════════════════════

def _render_details_table(df: pd.DataFrame, sales_df: pd.DataFrame) -> None:
    """Render the detailed issue ledger."""
    st.markdown("### 📋 Issue Ledger")

    # Guard: Check required columns exist
    if "date" not in df.columns:
        st.info("📊 Return data is loading... Date information not yet available.")
        return

    # Prepare display df
    display_df = df.copy()

    # Apply SKU mapping and item breakdown
    with st.spinner("Resolving Order Items..."):
        display_df["Item Breakdown"] = display_df.apply(
            lambda row: get_order_items_breakdown(row["order_id"], row["returned_items"], sales_df),
            axis=1
        )
    
    # Format the items list for display
    def format_item_list(items):
        if items is None or (hasattr(items, '__len__') and len(items) == 0):
            return "N/A"
        if not isinstance(items, list): return str(items)
        
        formatted = []
        for i in items:
            if isinstance(i, dict):
                name = i.get("name", "Unknown")
                sku = i.get("sku", "N/A")
                size = i.get("size", "N/A")
                qty = i.get("qty", 1)
                cat = i.get("category", "N/A")
            else:
                # Fallback for old string format
                name = str(i)
                sku = "N/A"; size = "N/A"; qty = 1; cat = "N/A"
            
            # Format: Name (SKU) [Size] xQty {Cat}
            line = f"**{name}** ({sku})"
            if size != "N/A": line += f" | Size: {size}"
            if qty > 1: line += f" | x{qty}"
            if cat != "General" and cat != "N/A": line += f" | {cat}"
            formatted.append(line)
            
        return "  \n".join(formatted)

    display_df["Returned Items (Details)"] = display_df["Item Breakdown"].apply(lambda x: format_item_list(x.get("returned", [])))
    display_df["Delivered Items (Details)"] = display_df.apply(
        lambda x: format_item_list(x["Item Breakdown"].get("delivered", [])) if x["issue_type"] == "Partial" else "-", 
        axis=1
    )

    display_cols = [
        "date", "order_id_raw", "issue_type", "return_reason",
        "Returned Items (Details)", "Delivered Items (Details)", 
        "customer_reason", "courier_reason",
        "fu_status", "inventory_updated", "partial_amount",
    ]
    # Filter only available columns that exist in our prepared display_df
    available_cols = [c for c in display_cols if c in display_df.columns]
    display_df = display_df[available_cols].copy()

    # Format for UI
    col_rename = {
        "date": "Date",
        "order_id_raw": "Order ID",
        "issue_type": "Type",
        "return_reason": "Reason",
        "customer_reason": "Customer Reason",
        "courier_reason": "Courier Reason",
        "fu_status": "Follow-Up",
        "inventory_updated": "Inventory",
        "partial_amount": "Partial ৳",
    }
    display_df = display_df.rename(columns=col_rename)

    if "Date" in display_df.columns:
        display_df["Date"] = display_df["Date"].dt.strftime("%Y-%m-%d")

    st.dataframe(
        display_df,
        width="stretch",
        height=500,
        column_config={
            "Partial ৳": st.column_config.NumberColumn(format="৳%.0f"),
        },
    )

    st.caption(f"Showing {len(display_df)} records")


# ═══════════════════════════════════════════════════════════════════
# EXPORT
# ═══════════════════════════════════════════════════════════════════

def _render_export(df: pd.DataFrame, metrics: dict) -> None:
    """Export returns data to Excel with summary sheet."""
    from io import BytesIO

    st.markdown("### 📤 Export Report")

    if st.button("📊 Export Returns Report", type="primary", width="content"):
        with st.spinner("Generating report..."):
            buffer = _generate_excel_report(df, metrics)
            st.download_button(
                label="⬇️ Download Excel Report",
                data=buffer,
                file_name=f"deen_returns_report_{datetime.now().strftime('%Y%m%d')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )


def _generate_excel_report(df: pd.DataFrame, metrics: dict) -> bytes:
    """Generate a multi-sheet Excel report."""
    from io import BytesIO

    try:
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        pass

    buffer = BytesIO()

    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        # ── Summary Sheet ──
        summary_data = {
            "Metric": [
                "Total Issues Tracked",
                "Total Returns (Paid + Non-Paid)",
                "  ├─ Paid Returns",
                "  └─ Non-Paid Returns",
                "Partial Orders",
                "Exchanges",
                "",
                "Return Rate (%)",
                "Partial Amounts (৳)",
            ],
            "Value": [
                metrics["total_issues"],
                metrics["return_count"],
                metrics.get("paid_return_count", 0),
                metrics.get("non_paid_return_count", 0),
                metrics["partial_count"],
                metrics["exchange_count"],
                "",
                metrics.get("return_rate", 0),
                metrics["partial_amounts"],
            ],
        }
        summary_df = pd.DataFrame(summary_data)
        summary_df.to_excel(writer, sheet_name="Summary", index=False)

        # ── Reason Breakdown ──
        reasons = metrics.get("reason_counts", {})
        if reasons:
            reason_df = pd.DataFrame([
                {"Reason": k, "Count": v} for k, v in reasons.items()
            ]).sort_values("Count", ascending=False)
            reason_df.to_excel(writer, sheet_name="Reason Analysis", index=False)

        # ── Detailed Data ──
        export_cols = [
            "date", "order_id_raw", "order_id", "issue_type", "return_reason",
            "product_details", "customer_reason", "courier_reason",
            "courier", "fu_status", "inventory_updated", "partial_amount",
        ]
        available = [c for c in export_cols if c in df.columns]
        detail_df = df[available].copy()
        if "date" in detail_df.columns:
            detail_df["date"] = detail_df["date"].dt.strftime("%Y-%m-%d")
        detail_df.to_excel(writer, sheet_name="Detailed Ledger", index=False)

        # ── Style sheets ──
        try:
            wb = writer.book
            header_fill = PatternFill(start_color="1a1a2e", end_color="1a1a2e", fill_type="solid")
            header_font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
            thin_border = Border(
                left=Side(style="thin"),
                right=Side(style="thin"),
                top=Side(style="thin"),
                bottom=Side(style="thin"),
            )

            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                for cell in ws[1]:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal="center")
                    cell.border = thin_border

                for col in ws.columns:
                    max_len = max(len(str(cell.value or "")) for cell in col)
                    ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 50)

                ws.freeze_panes = "A2"
        except Exception:
            pass

    return buffer.getvalue()
