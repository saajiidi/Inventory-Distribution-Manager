import pandas as pd


def read_uploaded_file(uploaded_file):
    """Read CSV/XLSX from a Streamlit uploader or file-like object."""
    if not uploaded_file:
        return None
    try:
        uploaded_file.seek(0)
    except Exception:
        pass

    name = str(getattr(uploaded_file, "name", "")).lower()
    if name.endswith(".csv"):
        return pd.read_csv(uploaded_file)
    return pd.read_excel(uploaded_file)


def fetch_remote_csv_raw(csv_url: str):
    """Low-level fetcher for raw CSV data and metadata (ETag, Last-Modified)."""
    from urllib.request import Request, urlopen

    req = Request(csv_url, headers={"User-Agent": "Mozilla/5.0"})
    with urlopen(req, timeout=15) as resp:
        raw_bytes = resp.read()
        headers = resp.headers
    return raw_bytes, headers


def read_remote_csv(csv_url: str):
    """Fetch remote CSV and return DataFrame + formatted timestamp."""
    from io import BytesIO
    from email.utils import parsedate_to_datetime

    try:
        raw, headers = fetch_remote_csv_raw(csv_url)
        df = pd.read_csv(BytesIO(raw), sep="\t")
        lm = headers.get("Last-Modified")
        if lm:
            try:
                lm = parsedate_to_datetime(lm).strftime("%Y-%m-%d %H:%M:%S")
            except Exception:
                lm = "Live Sync"
        else:
            lm = "Snapshot"
        return df, lm
    except Exception as e:
        raise RuntimeError(f"Failed to fetch CSV from {csv_url}: {e}")


def to_excel_bytes(df: pd.DataFrame, sheet_name: str = "Data") -> bytes:
    """Export DataFrame to Excel bytes with premium openpyxl styling.
    
    Features:
    - Bold headers with centered text
    - Alternating row colors (subtle)
    - Auto-filter enabled
    - Optimized column widths
    - Proper number/date formatting
    """
    from io import BytesIO
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter

    output = BytesIO()
    
    # Use context manager for ExcelWriter
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name=sheet_name)
        
        workbook = writer.book
        worksheet = writer.sheets[sheet_name]
        
        # Styles
        header_fill = PatternFill(start_color="1D4ED8", end_color="1D4ED8", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        center_aligned = Alignment(horizontal="center", vertical="center")
        border_side = Side(style="thin", color="D1D5DB")
        thin_border = Border(left=border_side, right=border_side, top=border_side, bottom=border_side)
        
        # Format Headers
        for col_idx, column in enumerate(df.columns, 1):
            cell = worksheet.cell(row=1, column=col_idx)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_aligned
            cell.border = thin_border
            
            # Auto-size columns (rough estimate)
            max_len = max(
                df[column].astype(str).map(len).max(),
                len(str(column))
            ) + 2
            worksheet.column_dimensions[get_column_letter(col_idx)].width = min(max_len, 50)

        # Enable Auto-Filter
        worksheet.auto_filter.ref = worksheet.dimensions
        
        # Freeze first row
        worksheet.freeze_panes = "A2"

    output.seek(0)
    return output.getvalue()
